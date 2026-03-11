from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from dotenv import load_dotenv
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

from translator import translate_text, translate_chunks, translate_segments, get_language_code
# from validator import validate_translation
from doc_handler import (
    extract_text_from_pdf,
    create_translated_pdf,
    extract_text_from_docx,
    extract_text_plain_from_docx,
    create_frozen_template_pdf,
)

app = FastAPI(title="IFU Translator API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5174",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:8080"
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================================
# PDF Endpoints
# ============================================================================

@app.post("/extract-pdf")
async def extract_pdf(file: UploadFile = File(...)):
    """Extract text from a PDF file."""
    pdf_bytes = await file.read()
    text = extract_text_from_pdf(pdf_bytes)
    return {"text": text, "filename": file.filename}


# ============================================================================
# DOCX Endpoints
# ============================================================================

@app.post("/extract-docx")
async def extract_docx(file: UploadFile = File(...)):
    """
    Extract structured segments from a DOCX file.
    Returns segments with id, type (h1/h2/h3/p/li/ol), and text.
    """
    docx_bytes = await file.read()
    segments = extract_text_from_docx(docx_bytes)
    return {"segments": segments, "filename": file.filename}


@app.post("/extract-docx-plain")
async def extract_docx_plain(file: UploadFile = File(...)):
    """Extract plain text from DOCX for translation."""
    docx_bytes = await file.read()
    text = extract_text_plain_from_docx(docx_bytes)
    return {"text": text, "filename": file.filename}


# ============================================================================
# Translation Endpoints
# ============================================================================

@app.post("/translate")
async def translate(
    text: str = Form(...),
    target_lang: str = Form("French")
):
    """
    Translate text to target language.
    Uses streaming to report progress.
    """
    async def generate():
        try:
            translated_chunks = []
            for chunk_text, step, total in translate_chunks(text, target_lang):
                translated_chunks.append(chunk_text)
                progress = round(step / total * 100)
                yield json.dumps({"type": "progress", "value": progress, "step": step, "total": total}) + "\n"
            yield json.dumps({"type": "done", "translation": " ".join(translated_chunks)}) + "\n"
        except Exception as e:
            yield json.dumps({"type": "error", "detail": str(e)}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")


@app.post("/translate-segments")
async def translate_segments_endpoint(
    file: UploadFile = File(...),
    target_lang: str = Form("French")
):
    """
    Translate DOCX segments to target language.
    This is the main endpoint for the IFU translation workflow.
    """
    docx_bytes = await file.read()
    
    try:
        # Extract segments from DOCX
        segments = extract_text_from_docx(docx_bytes)
        
        # Translate each segment
        translated_segments = []
        for result in translate_segments(segments, target_lang):
            translated_segments.append(result)
        
        return {
            "segments": translated_segments,
            "target_lang": target_lang,
            "source_lang": "English"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# PDF Generation Endpoints
# ============================================================================

@app.post("/export-pdf")
async def export_pdf(
    text: str = Form(...),
    filename: str = Form("translated"),
    segments_json: Optional[str] = Form(None)
):
    """
    Generate a PDF from translated text or segments with proper formatting.
    If segments_json is provided, uses segment types for formatting.
    """
    try:
        print(f"Generating PDF for text length: {len(text)}")
        
        # Check if we have segment information for proper formatting
        if segments_json:
            try:
                import json
                segments = json.loads(segments_json)
                if segments and len(segments) > 0:
                    # Use frozen template PDF with segment types
                    pdf_bytes = create_frozen_template_pdf(
                        original_file=b'',
                        segments=segments,
                        target_lang="",
                        doc_title=filename.replace(".pdf", "").replace("_", " ").title(),
                        doc_ref="DOC-001"
                    )
                    print(f"PDF generated with formatting, size: {len(pdf_bytes)} bytes")
                    
                    header = pdf_bytes[:8]
                    print(f"PDF header: {header}")
                    print(f"Is valid PDF: {header == b'%PDF-1.'}")
                    
                    headers = {
                        "Content-Disposition": f'attachment; filename="{filename}.pdf"',
                        "Content-Type": "application/pdf",
                    }
                    
                    return Response(
                        content=pdf_bytes,
                        media_type="application/pdf",
                        headers=headers,
                    )
            except json.JSONDecodeError as e:
                print(f"Failed to parse segments JSON: {e}")
        
        # Fallback to plain text PDF
        pdf_bytes = create_translated_pdf(text, filename)
        print(f"PDF generated, size: {len(pdf_bytes)} bytes")
        
        # Verify PDF header
        header = pdf_bytes[:8]
        print(f"PDF header: {header}")
        print(f"Is valid PDF: {header == b'%PDF-1.'}")
        
        # Create response with proper headers
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}.pdf"',
            "Content-Type": "application/pdf",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0"
        }
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers=headers,
        )
    except Exception as e:
        print(f"Export PDF error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/export-frozen-pdf")
async def export_frozen_pdf(
    file: UploadFile = File(...),
    segments: str = Form(...),  # JSON string of translated segments
    target_lang: str = Form("French"),
    doc_title: str = Form("IFU Document"),
    doc_ref: str = Form("IFU-001")
):
    """
    Generate a PDF using the frozen template approach.
    Preserves original DOCX formatting while replacing text with translations.
    """
    try:
        # Parse the segments JSON
        import ast
        translated_segments = json.loads(segments)
        
        # Get original DOCX bytes
        docx_bytes = await file.read()
        
        # Create frozen template PDF
        pdf_bytes = create_frozen_template_pdf(
            original_file=docx_bytes,
            segments=translated_segments,
            target_lang=target_lang,
            doc_title=doc_title,
            doc_ref=doc_ref
        )
        
        filename = f"{doc_title}_{target_lang.replace(' ', '_')}"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Validation Endpoint
# ============================================================================

class ValidateRequest(BaseModel):
    source: str
    translation: str
    reference: Optional[str] = None


@app.post("/validate")
async def validate(req: ValidateRequest):
    """Validate translation quality (placeholder - requires validator module)."""
    # Placeholder response since validator module is not implemented
    return {
        "score": 0.0,
        "issues": [],
        "message": "Validation not implemented. Install validator module for translation quality checks."
    }


# ============================================================================
# Excel Export Endpoint
# ============================================================================

class Correction(BaseModel):
    original: str
    mistranslated: str
    correct: str
    context: str


class ExcelExportRequest(BaseModel):
    corrections: List[Correction]
    document_name: Optional[str] = "translation"


@app.post("/export-excel")
async def export_excel(req: ExcelExportRequest):
    """Export translation corrections to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Translation Corrections"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_fill_even = PatternFill("solid", fgColor="EFF6FF")
    row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    critical_fill = PatternFill("solid", fgColor="FEF2F2")

    cell_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    headers = ["#", "Original (English)", "Mistranslated As", "Correct Translation", "Context / Notes"]
    col_widths = [5, 35, 35, 35, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28

    # Data rows
    critical_keywords = ["critical", "sterile", "implant", "contraindic", "warning", "caution", "dose", "dosage"]

    for i, c in enumerate(req.corrections, start=1):
        row = i + 1
        is_critical = any(kw in c.context.lower() or kw in c.original.lower() for kw in critical_keywords)
        fill = critical_fill if is_critical else (row_fill_even if i % 2 == 0 else row_fill_odd)

        values = [i, c.original, c.mistranslated, c.correct, c.context]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.alignment = cell_align
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.font = Font(name="Calibri", bold=True, color="64748B")
            elif col == 4:
                cell.font = Font(name="Calibri", color="15803D", bold=True)
            elif col == 3:
                cell.font = Font(name="Calibri", color="DC2626")

        ws.row_dimensions[row].height = 42

    # Summary row
    if req.corrections:
        summary_row = len(req.corrections) + 3
        ws.cell(row=summary_row, column=1, value=f"Total issues: {len(req.corrections)}")
        ws.cell(row=summary_row, column=1).font = Font(bold=True, color="1E3A5F")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{req.document_name}_corrections.xlsx"
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# Health Check
# ============================================================================

@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "IFU Translator API"}


@app.get("/test-pdf")
async def test_pdf():
    """Test endpoint that returns a simple valid PDF."""
    # Create a minimal valid PDF
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_text(fitz.Point(50, 50), "Test PDF - This is a test!", fontsize=16, fontname="helv")
    pdf_bytes = doc.tobytes()
    
    print(f"Test PDF generated, size: {len(pdf_bytes)}, header: {pdf_bytes[:10]}")
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="test.pdf"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)

